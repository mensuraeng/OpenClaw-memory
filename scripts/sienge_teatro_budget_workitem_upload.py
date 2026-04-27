#!/usr/bin/env python3
"""Automated Teatro Suzano budget upload path using Sienge workItemId + quantity.

This is the only public-API-compatible path found so far:
- preserve current Sienge sheet items by id
- add imported Excel items as sheet items with workItemId and quantity
- Sienge calculates prices from its cost database/service composition

It does NOT push Excel unitPrice directly; public docs do not expose that write endpoint.
Default is dry-run. Use --execute only after reviewing coverage/totals.
"""
from __future__ import annotations

import argparse
import base64
import json
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

WORKSPACE = Path('/root/.openclaw/workspace')
CONFIG = WORKSPACE / 'config/sienge-pcs.json'
DEFAULT_XLSX = Path('/root/.openclaw/media/inbound/TEATRO_SUZANO_-_Orçamento_SIENGE---0c16c0e6-dbdc-4682-b803-b058c09ef44f.xlsx')
RUNTIME = WORKSPACE / 'runtime/sienge/teatro-suzano'
BUILDING_ID = 1354
SHEET_ID = 1


def money(v: Any) -> float | None:
    if v in (None, ''):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace('R$', '').replace('.', '').replace(',', '.').strip())
    except Exception:
        return None


class Sienge:
    def __init__(self, config_path: Path = CONFIG):
        cfg = json.loads(config_path.read_text())
        self.base = f"https://api.sienge.com.br/{cfg['subdomain']}/public/api/v1"
        token = base64.b64encode(f"{cfg['username']}:{cfg['password']}".encode()).decode()
        self.auth = 'Basic ' + token

    def request(self, method: str, path: str, params: dict | None = None, body: Any | None = None, timeout: int = 90):
        query = '?' + urllib.parse.urlencode(params, doseq=True) if params else ''
        headers = {'Authorization': self.auth, 'Accept': 'application/json'}
        data = None
        if body is not None:
            headers['Content-Type'] = 'application/json'
            data = json.dumps(body, ensure_ascii=False).encode('utf-8')
        req = urllib.request.Request(f'{self.base}/{path}{query}', data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                text = resp.read().decode('utf-8', 'replace')
                return resp.status, text
        except urllib.error.HTTPError as exc:
            return exc.code, exc.read().decode('utf-8', 'replace')

    def get_json(self, path: str, params: dict | None = None):
        status, text = self.request('GET', path, params=params)
        if status == 429:
            raise RuntimeError('RATE_LIMIT_429')
        if status >= 400:
            raise RuntimeError(f'HTTP {status}: {text[:400]}')
        return json.loads(text)


def parse_excel(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Orçamento'] if 'Orçamento' in wb.sheetnames else wb.active
    current_section = ''
    section_order: list[str] = []
    items: list[dict] = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        item_col = row[1] if len(row) > 1 else None
        ref_col = row[2] if len(row) > 2 else None
        code_col = row[3] if len(row) > 3 else None
        desc_col = row[4] if len(row) > 4 else None
        unit_col = row[5] if len(row) > 5 else None
        qty = money(row[6] if len(row) > 6 else None)
        price = money(row[8] if len(row) > 8 else None)
        total = money(row[9] if len(row) > 9 else None)
        if isinstance(item_col, int) and isinstance(ref_col, str) and not code_col and not desc_col:
            current_section = ref_col.strip()
            if current_section not in section_order:
                section_order.append(current_section)
            continue
        if not code_col or not desc_col or qty is None or price is None:
            continue
        if not current_section:
            current_section = 'SEM SEÇÃO'
            if current_section not in section_order:
                section_order.append(current_section)
        items.append({
            'item': str(item_col).strip() if item_col else '',
            'ref': str(ref_col or '').strip(),
            'code': str(code_col).strip(),
            'section': current_section,
            'description': str(desc_col).strip(),
            'unit': str(unit_col or '').strip(),
            'quantity': qty,
            'excel_unit_price': price,
            'excel_total': total if total is not None else round(qty * price, 2),
        })
    return items, section_order


def paged(client: Sienge, path: str, params: dict | None = None, sleep_s: float = 0.2):
    params = dict(params or {})
    params.setdefault('limit', 200)
    offset = 0
    out = []
    while True:
        params['offset'] = offset
        data = client.get_json(path, params=params)
        rows = data.get('results', [])
        out.extend(rows)
        meta = data.get('resultSetMetadata', {})
        count = meta.get('count', len(out))
        if not rows or offset + len(rows) >= count:
            break
        offset += len(rows)
        time.sleep(sleep_s)
    return out


def build_workitem_index(client: Sienge, db_ids: list[int]):
    index: dict[str, list[dict]] = {}
    totals = {}
    for db in db_ids:
        rows = paged(client, f'cost-databases/{db}/work-items', sleep_s=0.35)
        totals[db] = len(rows)
        for row in rows:
            aux = str(row.get('auxiliaryCode') or '').strip()
            if aux:
                row['_db'] = db
                index.setdefault(aux, []).append(row)
    return index, totals


def choose_workitem(item: dict, candidates: list[dict]):
    if not candidates:
        return None
    ref = item['ref'].upper()
    if ref.startswith('SINAPI'):
        pref = [3, 2, 1]
    elif ref in {'CDHU', 'EDIF', 'FDE'}:
        pref = [2, 1, 3]
    else:
        pref = [2, 3, 1]
    for db in pref:
        rows = [c for c in candidates if c.get('_db') == db]
        if rows:
            return rows[0]
    return candidates[0]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--file', default=str(DEFAULT_XLSX))
    ap.add_argument('--building-id', type=int, default=BUILDING_ID)
    ap.add_argument('--sheet-id', type=int, default=SHEET_ID)
    ap.add_argument('--db', type=int, action='append', default=[1, 2, 3], help='Cost database IDs to scan')
    ap.add_argument('--execute', action='store_true', help='Actually PUT sheet items')
    args = ap.parse_args()

    RUNTIME.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d-%H%M%S')
    client = Sienge()
    items, sections = parse_excel(Path(args.file))
    before = client.get_json(f'building-cost-estimations/{args.building_id}/sheets/{args.sheet_id}/items')
    current_items = before.get('results', [])

    index, db_totals = build_workitem_index(client, args.db)
    matched = []
    missing = []
    for item in items:
        wi = choose_workitem(item, index.get(item['code'], []))
        if wi:
            matched.append({**item, 'workItemId': wi['id'], 'sienge_description': wi.get('description'), 'sienge_unit_price': wi.get('unitPrice'), 'sienge_db': wi.get('_db')})
        else:
            missing.append(item)

    payload = []
    root = next((x for x in current_items if str(x.get('id')) == '3'), current_items[0] if current_items else None)
    payload.append({'id': str(root.get('id') if root else '3'), 'description': root.get('description', 'OBRA GERAL') if root else 'OBRA GERAL', 'workItemId': root.get('workItemId') if root else None, 'quantity': root.get('quantity') if root else None, 'level': 1})
    for old in current_items:
        if str(old.get('id')) == '3':
            continue
        payload.append({'id': str(old.get('id')), 'description': old.get('description'), 'workItemId': old.get('workItemId'), 'quantity': old.get('quantity'), 'level': 2})
    by_section: dict[str, list[dict]] = {}
    for item in matched:
        by_section.setdefault(item['section'], []).append(item)
    for section in sections:
        rows = by_section.get(section, [])
        if not rows:
            continue
        payload.append({'description': f'IMPORTADO PLANILHA - {section}', 'workItemId': None, 'quantity': None, 'level': 2})
        for item in rows:
            payload.append({'description': f"[{item['item']}] {item['description']}"[:500], 'workItemId': item['workItemId'], 'quantity': item['quantity'], 'level': 3})

    summary = {
        'mode': 'EXECUTE' if args.execute else 'DRY_RUN',
        'excel_items': len(items),
        'matched_by_auxiliary_code': len(matched),
        'missing_workitem': len(missing),
        'current_items_preserved': len(current_items),
        'payload_entries': len(payload),
        'excel_total': round(sum(x['excel_total'] for x in items), 2),
        'sienge_reference_total_for_matched': round(sum((x.get('sienge_unit_price') or 0) * x['quantity'] for x in matched), 2),
        'cost_database_rows_scanned': db_totals,
        'missing': missing,
    }
    (RUNTIME / f'workitem-upload-summary-{ts}.json').write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
    (RUNTIME / f'workitem-upload-payload-{ts}.json').write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    print(json.dumps({k: summary[k] for k in summary if k != 'missing'}, ensure_ascii=False, indent=2))
    if missing:
        print('MISSING_FIRST=', json.dumps(missing[:20], ensure_ascii=False, indent=2))
    if not args.execute:
        print('DRY_RUN_ONLY. Use --execute to PUT this structure.')
        return
    if missing:
        raise SystemExit('Refusing execute with missing workItemId mappings')
    status, text = client.request('PUT', f'building-cost-estimations/{args.building_id}/sheets/{args.sheet_id}/items', body=payload, timeout=180)
    (RUNTIME / f'workitem-upload-response-{ts}.txt').write_text(text, encoding='utf-8')
    print(f'PUT_STATUS={status}')
    print(text[:1200])
    if status not in (200, 201, 204):
        raise SystemExit(2)
    after = client.get_json(f'building-cost-estimations/{args.building_id}/sheets/{args.sheet_id}/items', params={'limit': 200})
    (RUNTIME / f'workitem-upload-after-{ts}.json').write_text(json.dumps(after, ensure_ascii=False, indent=2), encoding='utf-8')
    print('AFTER_COUNT=', len(after.get('results', [])))


if __name__ == '__main__':
    main()
