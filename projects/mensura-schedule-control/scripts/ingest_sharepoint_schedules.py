#!/usr/bin/env python3
"""Ingest downloaded SharePoint Arquivos Auxiliares schedule tables into Supabase.

Reads runtime/sharepoint/manifest.json and workbook_inspection.json, extracts named tables
containing MSchedule/MasterSchedule from local xlsx files, and inserts:
- schedule.projects
- raw.import_jobs
- raw.source_rows
- schedule.schedule_versions
- schedule.activity_identities
- schedule.activity_versions

This importer is intentionally conservative: it does not infer dependencies yet.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any

import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

ROOT = Path(__file__).resolve().parents[1]
MANIFEST = ROOT / "runtime/sharepoint/manifest.json"
INSPECTION = ROOT / "runtime/sharepoint/workbook_inspection.json"

SCHEDULE_TABLE_RE = re.compile(r"schedule", re.I)

COLUMN_ALIASES = {
    "activity_code": ["id", "activity id", "código", "codigo"],
    "activity_name": ["tarefas", "nome da tarefa", "atividade", "nome", "nome de origem"],
    "wbs_1": ["os. - 1º nível", "p.s. - 1º nível", "ps. - 1º nível", "texto12"],
    "wbs_2": ["ps. - 2º nível", "p.s. - 2º nível2", "ps. - 2º nìvel", "texto9"],
    "wbs_3": ["ps. - 3º nível", "p.s. - 3º nível2", "ps. - 3º nìvel", "texto10"],
    "discipline": ["etapa"],
    "location": ["pavto.", "pavto", "torre", "obras", "ambiente"],
    "planned_duration_days": ["duração", "duracao"],
    "remaining_duration_days": ["duração restante", "duracao restante"],
    "baseline_start": ["início da linha de base", "inicio da linha de base", "início da linha de base1", "início da linha de base (meta)"],
    "baseline_finish": ["término da linha de base", "termino da linha de base", "término da linha de base1", "término da linha de base (meta)"],
    "current_start": ["início", "inicio"],
    "current_finish": ["término", "termino"],
    "actual_start": ["início real", "inicio real"],
    "actual_finish": ["término real", "termino real"],
    "percent_complete": ["% concluída", "% concluido", "% concluído", "% executado", "avancfisico"],
    "physical_percent_complete": ["% concluída", "% concluido", "% concluído", "avancfisico"],
    "is_critical": ["crítica", "critica"],
    "status": ["status atividades", "status das tarefas", "status das tarefas - completo"],
    "responsible_party": ["responsáveis", "responsaveis"],
    "total_float_days": ["margem de atraso total", "folga total", "total float"],
}


def norm(s: Any) -> str:
    s = "" if s is None else str(s).strip().lower()
    s = s.replace("ì", "í")
    s = re.sub(r"\s+", " ", s)
    return s


def slug_code(name: str) -> str:
    base = Path(name).stem
    base = re.sub(r"^\d+-", "", base)
    base = base.replace("Arquivos Auxiliares - ", "").replace(" - Arquivos Auxiliares", "")
    base = base.replace("MasterSchedule", "Master Schedule")
    s = re.sub(r"[^A-Za-z0-9]+", "_", base).strip("_").upper()
    return s[:80] or "PROJETO"


def parse_date(v: Any):
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    txt = str(v).strip()
    if not txt:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(txt[:19], fmt).date()
        except Exception:
            pass
    return None


def parse_num(v: Any):
    if v in (None, ""):
        return None
    if isinstance(v, bool):
        return 1 if v else 0
    if isinstance(v, (int, float)):
        # Excel percent sometimes appears as 0.75; preserve duration-like as is elsewhere.
        return float(v)
    txt = str(v).strip().replace("%", "").replace(".", "").replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", txt)
    return float(m.group()) if m else None


def parse_bool(v: Any):
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = norm(v)
    if s in ("sim", "yes", "true", "verdadeiro", "1", "crítica", "critica"):
        return True
    if s in ("não", "nao", "no", "false", "falso", "0"):
        return False
    return None


def table_rows(path: Path, sheet: str, ref: str):
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb[sheet]
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    rows = list(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))
    if not rows:
        return [], []
    headers = [str(x).strip() if x is not None else f"col_{i+1}" for i, x in enumerate(rows[0])]
    out = []
    for idx, row in enumerate(rows[1:], start=2):
        d = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        if any(v not in (None, "") for v in d.values()):
            d["__source_row_number"] = idx
            out.append(d)
    return headers, out


def map_row(raw: dict[str, Any], headers: list[str]) -> dict[str, Any]:
    nmap = {norm(h): h for h in headers}
    def get_field(key):
        for alias in COLUMN_ALIASES.get(key, []):
            h = nmap.get(norm(alias))
            if h is not None:
                return raw.get(h)
        return None
    wbs_parts = [get_field("wbs_1"), get_field("wbs_2"), get_field("wbs_3")]
    wbs_code = " / ".join(str(x).strip() for x in wbs_parts if x not in (None, "")) or None
    act_name = get_field("activity_name")
    if act_name is None or str(act_name).strip() == "":
        act_name = get_field("activity_code")
    pc = parse_num(get_field("percent_complete"))
    if pc is not None and 0 <= pc <= 1:
        pc *= 100
    ppc = parse_num(get_field("physical_percent_complete"))
    if ppc is not None and 0 <= ppc <= 1:
        ppc *= 100
    return {
        "activity_code": str(get_field("activity_code") or "").strip() or None,
        "activity_name": str(act_name or "").strip() or None,
        "wbs_code": wbs_code,
        "discipline": str(get_field("discipline") or "").strip() or None,
        "location": str(get_field("location") or "").strip() or None,
        "responsible_party": str(get_field("responsible_party") or "").strip() or None,
        "planned_duration_days": parse_num(get_field("planned_duration_days")),
        "remaining_duration_days": parse_num(get_field("remaining_duration_days")),
        "baseline_start": parse_date(get_field("baseline_start")),
        "baseline_finish": parse_date(get_field("baseline_finish")),
        "current_start": parse_date(get_field("current_start")),
        "current_finish": parse_date(get_field("current_finish")),
        "actual_start": parse_date(get_field("actual_start")),
        "actual_finish": parse_date(get_field("actual_finish")),
        "percent_complete": pc,
        "physical_percent_complete": ppc,
        "is_critical": parse_bool(get_field("is_critical")),
        "total_float_days": parse_num(get_field("total_float_days")),
        "status": str(get_field("status") or "").strip() or None,
        "raw_payload": raw,
    }


def connect():
    dsn = os.environ.get("MENSURA_SCHEDULE_DATABASE_URL") or os.environ.get("DATABASE_URL")
    if dsn:
        return psycopg2.connect(dsn)
    password = os.environ.get("MENSURA_SCHEDULE_DB_PASSWORD")
    host = os.environ.get("MENSURA_SCHEDULE_DB_HOST", "db.ckmuyvbacgdidmiccvif.supabase.co")
    return psycopg2.connect(host=host, port=5432, dbname="postgres", user="postgres", password=password, sslmode="require")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--limit", type=int, default=0)
    args = ap.parse_args()

    manifest = json.loads(MANIFEST.read_text())
    inspections = {x.get("local_path"): x for x in json.loads(INSPECTION.read_text())}
    candidates = []
    for m in manifest:
        lp = m.get("local_path")
        if not lp:
            continue
        ins = inspections.get(lp, {})
        for t in ins.get("tables", []):
            if SCHEDULE_TABLE_RE.search(t.get("table", "")):
                candidates.append((m, t))
    if args.limit:
        candidates = candidates[: args.limit]

    if args.dry_run:
        print(f"candidates={len(candidates)}")
    summary = []
    if args.dry_run:
        for m,t in candidates:
            headers, rows = table_rows(Path(m["local_path"]), t["sheet"], t["ref"])
            mapped = [map_row(r, headers) for r in rows]
            valid = [r for r in mapped if r.get("activity_name")]
            print(m["name"], t["table"], t["ref"], "rows", len(rows), "valid", len(valid))
        return

    conn = connect()
    conn.autocommit = False
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            for m,t in candidates:
                path = Path(m["local_path"])
                headers, rows = table_rows(path, t["sheet"], t["ref"])
                mapped = [map_row(r, headers) for r in rows]
                mapped = [r for r in mapped if r.get("activity_name")]
                if not mapped:
                    continue
                code = slug_code(m["name"])
                project_name = re.sub(r"^\d+-", "", Path(m["name"]).stem)
                cur.execute("""
                    insert into schedule.projects (code, name, company, status, metadata)
                    values (%s,%s,'Mensura','active',%s)
                    on conflict (code) do update set name=excluded.name, updated_at=now()
                    returning id
                """, (code, project_name, json.dumps({"source":"sharepoint_search","webUrl":m.get("webUrl")})))
                project_id = cur.fetchone()["id"]
                file_hash = hashlib.sha256(path.read_bytes()).hexdigest()
                data_date = (m.get("lastModified") or "")[:10] or None
                version_label = f"sharepoint-{data_date or datetime.utcnow().date()}-{t['table']}"
                cur.execute("""
                    select id from schedule.schedule_versions
                    where project_id = %s and version_label = %s
                """, (project_id, version_label))
                existing_version = cur.fetchone()
                if existing_version:
                    summary.append({"project":code,"file":m["name"],"table":t["table"],"rows":len(mapped),"status":"skipped_existing_version"})
                    continue
                cur.execute("""
                    insert into raw.import_jobs (project_id, source_tool, source_filename, source_uri, file_hash, imported_by, status, data_date, raw_metadata)
                    values (%s,'excel',%s,%s,%s,'flavia','validated',%s,%s)
                    returning id
                """, (project_id, m["name"], m.get("webUrl"), file_hash, data_date, json.dumps({"local_path":str(path),"table":t,"size":m.get("size"),"lastModified":m.get("lastModified")})))
                import_job_id = cur.fetchone()["id"]
                cur.execute("""
                    insert into schedule.schedule_versions (project_id, import_job_id, version_label, data_date, source_tool, notes, metadata)
                    values (%s,%s,%s,%s,'excel',%s,%s)
                    returning id
                """, (project_id, import_job_id, version_label, data_date or datetime.utcnow().date(), f"Imported from {m['name']} table {t['table']}", json.dumps({"sheet":t["sheet"],"ref":t["ref"],"headers":headers})))
                version_id = cur.fetchone()["id"]
                source_rows = []
                for r in rows:
                    source_rows.append((import_job_id, r.get("__source_row_number"), str(r.get("Id") or r.get("ID") or r.get("id") or "") or None, str(r.get("Id") or r.get("ID") or "") or None, json.dumps(r, default=str)))
                psycopg2.extras.execute_values(cur, """
                    insert into raw.source_rows (import_job_id, source_row_number, source_activity_uid, activity_code, raw_payload)
                    values %s
                    on conflict (import_job_id, source_row_number) do nothing
                """, source_rows, page_size=1000)
                identity_rows = []
                version_rows = []
                for idx, r in enumerate(mapped, start=1):
                    src_uid = r["activity_code"] or f"row-{idx}"
                    normalized = r["activity_name"][:500]
                    identity_rows.append((project_id, src_uid, r["activity_code"], normalized, r["wbs_code"]))
                    version_rows.append((
                        src_uid, r["activity_code"], r["activity_name"], r["wbs_code"], r["discipline"], r["location"],
                        r["planned_duration_days"], r["remaining_duration_days"], r["baseline_start"], r["baseline_finish"],
                        r["current_start"], r["current_finish"], r["actual_start"], r["actual_finish"], r["percent_complete"],
                        r["physical_percent_complete"], r["is_critical"], r["total_float_days"], r["status"], json.dumps(r["raw_payload"], default=str)
                    ))
                psycopg2.extras.execute_values(cur, """
                    insert into schedule.activity_identities (project_id, source_activity_uid, activity_code, normalized_name, wbs_code, identity_confidence)
                    values %s
                    on conflict (project_id, source_activity_uid) do update
                    set normalized_name=excluded.normalized_name,
                        activity_code=excluded.activity_code,
                        wbs_code=excluded.wbs_code
                """, identity_rows, template="(%s,%s,%s,%s,%s,'high')", page_size=1000)
                cur.execute("""
                    create temp table if not exists tmp_activity_versions_import (
                        source_activity_uid text,
                        activity_code text,
                        activity_name text,
                        wbs_code text,
                        discipline text,
                        location text,
                        planned_duration_days numeric,
                        remaining_duration_days numeric,
                        baseline_start date,
                        baseline_finish date,
                        current_start date,
                        current_finish date,
                        actual_start date,
                        actual_finish date,
                        percent_complete numeric,
                        physical_percent_complete numeric,
                        is_critical boolean,
                        total_float_days numeric,
                        status text,
                        raw_payload jsonb
                    ) on commit drop
                """)
                cur.execute("truncate tmp_activity_versions_import")
                psycopg2.extras.execute_values(cur, """
                    insert into tmp_activity_versions_import (
                        source_activity_uid, activity_code, activity_name, wbs_code, discipline, location,
                        planned_duration_days, remaining_duration_days, baseline_start, baseline_finish, current_start, current_finish,
                        actual_start, actual_finish, percent_complete, physical_percent_complete, is_critical, total_float_days, status, raw_payload
                    ) values %s
                """, version_rows, page_size=1000)
                cur.execute("""
                    insert into schedule.activity_versions (
                        schedule_version_id, activity_identity_id, activity_code, activity_name, wbs_code, discipline, location,
                        planned_duration_days, remaining_duration_days, baseline_start, baseline_finish, current_start, current_finish, actual_start, actual_finish,
                        percent_complete, physical_percent_complete, is_critical, total_float_days, status, raw_payload
                    )
                    select
                        %s, ai.id, tmp.activity_code, tmp.activity_name, tmp.wbs_code, tmp.discipline, tmp.location,
                        tmp.planned_duration_days, tmp.remaining_duration_days, tmp.baseline_start, tmp.baseline_finish,
                        tmp.current_start, tmp.current_finish, tmp.actual_start, tmp.actual_finish,
                        tmp.percent_complete, tmp.physical_percent_complete, tmp.is_critical, tmp.total_float_days, tmp.status, tmp.raw_payload
                    from tmp_activity_versions_import tmp
                    join schedule.activity_identities ai
                      on ai.project_id = %s
                     and ai.source_activity_uid = tmp.source_activity_uid
                    on conflict (schedule_version_id, activity_identity_id) do nothing
                """, (version_id, project_id))
                summary.append({"project":code,"file":m["name"],"table":t["table"],"rows":len(mapped)})
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    print(json.dumps(summary, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    main()
