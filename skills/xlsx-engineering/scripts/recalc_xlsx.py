#!/usr/bin/env python3
"""Recalculate XLSX/XLSM with LibreOffice headless and scan for Excel formula errors."""
import json, os, shutil, socket, subprocess, sys, tempfile, textwrap, time
from pathlib import Path
from openpyxl import load_workbook

ERRORS = ('#NULL!', '#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#NUM!', '#N/A', '#GETTING_DATA')

def scan(path):
    wb_formula = load_workbook(path, data_only=False, read_only=False)
    wb_values = load_workbook(path, data_only=True, read_only=False)
    errors, formulas = [], 0
    for s in wb_formula.sheetnames:
        ws_f, ws_v = wb_formula[s], wb_values[s]
        for row in ws_f.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith('='):
                    formulas += 1
                v = ws_v[c.coordinate].value
                if isinstance(v, str) and v.startswith(ERRORS):
                    errors.append({'sheet': s, 'cell': c.coordinate, 'value': v, 'formula': c.value})
    return formulas, errors

def free_port():
    s = socket.socket(); s.bind(('127.0.0.1', 0)); p = s.getsockname()[1]; s.close(); return p

def recalc_with_uno(path, timeout):
    soffice = shutil.which('soffice') or shutil.which('libreoffice')
    if not soffice:
        raise RuntimeError('LibreOffice/soffice not found')
    sys_py = '/usr/bin/python3' if Path('/usr/bin/python3').exists() else sys.executable
    port = free_port()
    with tempfile.TemporaryDirectory(prefix='lo-recalc-') as td:
        profile = Path(td) / 'profile'
        accept = f"socket,host=127.0.0.1,port={port};urp;StarOffice.ComponentContext"
        proc = subprocess.Popen([soffice, '--headless', '--nologo', '--nofirststartwizard', '--nodefault', f'-env:UserInstallation=file://{profile}', f'--accept={accept}'], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        try:
            deadline = time.time() + timeout
            while time.time() < deadline:
                try:
                    with socket.create_connection(('127.0.0.1', port), timeout=0.5): break
                except OSError: time.sleep(0.2)
            else:
                raise RuntimeError('LibreOffice listener did not start')
            helper = Path(td) / 'uno_recalc.py'
            helper.write_text(textwrap.dedent(f'''
                import pathlib, uno, time
                from com.sun.star.beans import PropertyValue
                local_ctx = uno.getComponentContext()
                resolver = local_ctx.ServiceManager.createInstanceWithContext('com.sun.star.bridge.UnoUrlResolver', local_ctx)
                ctx = resolver.resolve('uno:socket,host=127.0.0.1,port={port};urp;StarOffice.ComponentContext')
                smgr = ctx.ServiceManager
                desktop = smgr.createInstanceWithContext('com.sun.star.frame.Desktop', ctx)
                def prop(name, value):
                    p = PropertyValue(); p.Name = name; p.Value = value; return p
                url = pathlib.Path(r"{str(path)}").resolve().as_uri()
                doc = desktop.loadComponentFromURL(url, '_blank', 0, (prop('Hidden', True), prop('UpdateDocMode', 3)))
                if doc is None: raise RuntimeError('failed to open workbook')
                try:
                    doc.calculateAll()
                    doc.store()
                finally:
                    doc.close(True)
            '''), encoding='utf-8')
            res = subprocess.run([sys_py, str(helper)], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=timeout)
            if res.returncode != 0:
                raise RuntimeError(res.stderr.strip() or res.stdout.strip() or 'UNO recalc failed')
        finally:
            proc.terminate()
            try: proc.wait(timeout=5)
            except subprocess.TimeoutExpired: proc.kill()

def main():
    if len(sys.argv) < 2:
        print('usage: recalc_xlsx.py file.xlsx [timeout_seconds]', file=sys.stderr); return 2
    path = Path(sys.argv[1]).resolve(); timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 60
    if not path.exists():
        print(json.dumps({'status':'error','message':'file not found','file':str(path)})); return 1
    before_formulas, before_errors = scan(path)
    try:
        recalc_with_uno(path, timeout)
    except Exception as e:
        print(json.dumps({'status':'error','message':str(e),'file':str(path)}, ensure_ascii=False, indent=2)); return 1
    formulas, errors = scan(path)
    summary = {}
    for e in errors: summary[e['value']] = summary.get(e['value'], 0) + 1
    status = 'success' if not errors else 'errors_found'
    print(json.dumps({'status':status,'file':str(path),'total_formulas':formulas,'total_errors':len(errors),'error_summary':summary,'errors':errors[:50],'before_total_formulas':before_formulas,'before_total_errors':len(before_errors)}, ensure_ascii=False, indent=2))
    return 0 if not errors else 3

if __name__ == '__main__':
    raise SystemExit(main())
